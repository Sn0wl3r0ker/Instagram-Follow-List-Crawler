import requests, pickle
import re
from config import username,password,headers,url,ajax_url,p_url,path
from datetime import datetime
from openpyxl import Workbook, load_workbook
import compare, urlToPic, sort
import os, sys, time
import platform


def create_folder(path):
    if not os.path.isdir(path):
        os.makedirs(path)

def ask_excel(ask_option):
    flag = []
    yes_list = ['y','Y','yes']
    no_list = ['n','N','no','']
    while flag not in yes_list and flag not in no_list:
        flag = input(f'Do u want {ask_option}? y/n [n]: ')
        if(flag not in yes_list and flag not in no_list):
            print('plz enter y or n or ENTER!!!')
    # print(f'flag={flag}')
    if(flag in yes_list):
        return 1
    elif(flag in no_list):
        return 0
def ask_url_to_pic():
    ask3 = []
    yes_list = ['y','Y','yes','']
    no_list = ['n','N','no']
    while ask3 not in yes_list and ask3 not in no_list:
        ask3 = input(f'Do u want to transfer profile pic url to pics? It might *take more time*![y/n] [y]: ')
        if ask3 in yes_list:
            print(f'start getting pic files!!!')
            return 1
        elif ask3 in no_list:
            return 0
        else:
            print(f'plz enter y or n or empty to default!')

def do_excel(path,filename,root_json):              # 跑生成excel
    wb = Workbook()
    ws = wb.active
    title = ['username', 'full_name', 'profile_pic_url', 'Profile_Pic']
    ws.append(title)
    for users in root_json['users']:
        id = []
        id.append('@'+users['username'])
        id.append(users['full_name'])
        id.append(users['profile_pic_url']+'.jpg')
        ws.append(id)
    wb.save(path+filename+f'.xlsx')
    

def do_excel_mac(path,filename,root_json):
    if ask_url_to_pic() == 1:
        urlToPic.urlToPic.urlToPicinit(filename)
        time.sleep(3)
        print(f'waiting 3 seconds!')
        wb = load_workbook(os.path.abspath(path+filename+'pic.xlsm'), read_only=False, keep_vba=True)
        ws = wb.active
        title = ['username', 'full_name', 'profile_pic_url', 'Profile_Pic']
        ws.append(title)
        for users in root_json['users']:
            id = []
            id.append('@'+users['username'])
            id.append(users['full_name'])
            id.append(users['profile_pic_url']+'.jpg')
            ws.append(id)
        wb.save(path+filename+f'pic.xlsm')
        time.sleep(2)
        urlToPic.urlToPic.urlToPicMac(filename)
    else:
        do_excel(path,filename,root_json)

def do_txt(path,filename,root_json):                # 跑生成txt 
    i=1
    with open(path+filename+f'.txt', 'w+',encoding='utf-8') as f:
        for users in root_json['users']:
            # id = (f'{i}','@'+users['username'], users['full_name'])
            # print(type(users))
            # print(f'!!!{users}!!!')
            id = (f'@'+users['username'], users['full_name'])
            i+=1
        # reresponse = response.text.replace('\\u0026','&')
            f.write(str(id)+'\n')
        f.write(f'Total: {i-1} records!\n')
    print(f'Got {i-1} records!!!')
    flag = ask_excel('Do u need sort data for compare?[y/n]: ')
    if flag == 1:
        try:
            print('Sorting files!')
            sort.sorted_to_compare(path+filename+f'.txt')
        except UnicodeDecodeError as error:
            print('Sort ERROR~ plz try to use sort.py manually!!!')
            print(f'ERROR: {error}')
            sys.exit()

# def do_txt(path,filename,root_json,print_flag):                # 跑生成txt 
#     i=1
#     with open(path+filename+f'.txt', 'a+',encoding='utf-8') as f:
#         for users in root_json['users']:
#             # id = (f'{i}','@'+users['username'], users['full_name'])
#             id = (f'@'+users['username'], users['full_name'])
#             i+=1
#         # reresponse = response.text.replace('\\u0026','&')
#             f.write(str(id)+'\n')
#     if print_flag == 1:
#         with open(path+filename+f'.txt', 'r') as f:
#             # for count_line, line in enumerate(f):
#             #     pass
#             count_line = len(f.readlines())
#             print(f'Got {count_line} records!!!')
#         with open(path+filename+f'.txt', 'a+',encoding='utf-8') as f:
#             f.write(f'Total: {count_line} records!')


def main():
    system = platform.system()
    date = datetime.now().strftime("%Y%m%d-%H%M")
    timestamp = int(datetime.now().timestamp())
    payload = {
        'username': f'{username}',
        'enc_password': f'#PWD_INSTAGRAM_BROWSER:0:{timestamp}:{password}',
        'queryParams': {},
        'optIntoOneTap': 'false',
        'stopDeletionNonce': '',
        'trustedDeviceRecords': {}
    }
    # print(payload)
    print(f'If target is private account, you have to follow it first!!!')
    while True:
        uid = str(input('Enter id: '))
        if (uid == ''):
            print(f'Do not leave blank!!!')
        else:
            break
    while True:
        opt_list = {'':'following','1':'following','2':'followers','following':'following','followers':'followers'}
        option = str(input('following[1]/followers[2] [1]: '))
        if(option in opt_list):
            option = opt_list[option]
            break
        else:
            print(f'enter 1 or 2 or following or followers!!!')
    try:
        fcount = int(input('Enter target\'s max num of following/followers : '))
    except ValueError:
        print(f'plz enter the correct number!!!')
        sys.exit()

    opt_title = {
        'following': 'fwi',
        'followers': 'fwr',}
    filename = f'{uid}{date}{opt_title[option]}'
    ask = ask_excel('Excel file(will have profile pic)')
    # print(ask)
    with requests.session() as session:
        if not os.path.exists(f'{path}{username}session.pkl'):
            print('Getting sessions')      #session = requests.sess.....
            res = session.get(url)
            # with open('res.txt','a+', encoding='utf-8') as f:
            #     f.write(res.text)
            csrf = re.findall(r"csrf_token\":\"(.*?)\"}",res.text)[0]
            print(f'csrf={csrf}')
            # jazoest = re.findall(r"jazoest=(.*?)\"",res.text)[0]
            # print(f'jazoest={jazoest}')
            cookies = res.cookies                   #res獲取第一次cookie和csrf
            cookies['csrftoken'] = csrf
            headers['x-csrftoken'] = csrf
            # print("\n",headers,"\n")
            r = session.post(ajax_url, data=payload, headers=headers, cookies=cookies)
            print(r.status_code)
            with open(f'{path}{username}session.pkl', 'wb') as f:
                pickle.dump(session.cookies, f)        #用現有cookie和csrf token 去取得登入的session
            headers['Referer'] = f'https://www.instagram.com/{uid}/following/'
        # print(req2.text)
        else:
            print('Reloading sessions and updating cookies')
            print('If can\'t get all the records, plz delete session.pkl file and try again later!')
            headers['Referer'] = f'https://www.instagram.com/{uid}/following/'
            with open(f'{path}{username}session.pkl', 'rb') as f:
                cookies = session.cookies.update(pickle.load(f))
                headers['x-csrftoken'] = session.cookies['csrftoken']
                # print("\n",session.cookies)
                # print("\n",headers)
        fsi=session.get(p_url+uid,cookies=cookies,headers=headers)
        print(f'status: {fsi.status_code}')
        # with open('fsi.txt','a+', encoding='utf-8') as f:
        #         f.write(fsi.text)
        # print("\n",fsi.text)
        
        try:
            # print(str(re.findall(r"id\":\"(.*?)\"",fsi.text)))
            friendid = str(re.findall(r"\"id\":\"(.*?)\"",fsi.text)[0])
            checkid = str(re.findall(r"\"id\":\"(.*?)\"",fsi.text)[-1])
            if(friendid == '236' or friendid == None or checkid == '236'):
                raise Exception
            print(f"userid:{friendid}")
        except:
            # os.remove(f'{path}{username}session.pkl')
            print(f'error while checking userid')
            print(f'1.plz check the target username(no @)!!!')
            print(f'2.Make sure u set the right USERNAME and PASSWORD in *config.py* file!!!')
            print(f'3.your account might block by instagram server, plz try again later or change your ip!!')
            sys.exit()

        # url的後輟 可以像翻頁一樣去增加再爬取 或是直接爆max來爬取
        # params = {
        # 'count': fcount,
        # 'max_id': '200'}
        # response = session.get(f'https://i.instagram.com/api/v1/friendships/{friendid}/{option}/', params=params, cookies=cookies, headers=headers, timeout=300)
        # print_flag=0
        # if fcount % 200!=0:
        #     ffcount = int((fcount//200+1)*200)
        # else:
        #     ffcount = fcount
        # print(f'ffcount={ffcount}')
        c = 0
        while c <= fcount:
            # print(f'c={c}')
            if c == 0:
                response = session.get(f'https://i.instagram.com/api/v1/friendships/{friendid}/{option}/?count=100&search_surface=follow_list_page', cookies=cookies, headers=headers)
            else:
                response = session.get(f'https://i.instagram.com/api/v1/friendships/{friendid}/{option}/?count=100&max_id={max_id}&search_surface=follow_list_page', cookies=cookies, headers=headers)
            time.sleep(3)
            # with open('debug.txt','a+', encoding='utf-8') as f:
            #     f.write(response.text)
            try:
                if c == 0:
                    root_json = dict(response.json())
                    # print(f'FIRST:{root_json}')
                else:
                    main_json = dict(response.json())
                    root_json['users']+=(main_json['users'])
                    # print(f'SECOND:{root_json}')
            except requests.exceptions.JSONDecodeError as jsonError:
                print(f'Error when processing json file: {jsonError}')
                print(f'1.Make sure u set the right USERNAME and PASSWORD in *config.py* file!!!')
                print(f'2.your account might block by instagram server, plz try again later or change your ip!!')
                sys.exit()
            c+=100
            print(f'getting {c-100}~{c} records!')
            try:
                max_id = str(re.findall(r"\"next_max_id\":\"(.*?)\"",response.text)[-1])
                print(f'next_max_id={max_id}')
            except IndexError:
               finish = str(re.findall(r"\"has_more\":(.*?),",response.text)[-1])
               if(finish == 'false'):
                   break
            # show_many = session.post(f'https://i.instagram.com/api/v1/friendships/show_many/', cookies=cookies, headers=headers)
            # print(f'show_many:{show_many.status_code}')
        # with open(path+'root_json.txt', 'w+', encoding='utf-8') as jf:
        #     jf.write(str(root_json))
            # print(f'x:{c}')
            # if c == ffcount:
            #     print_flag = 1
        do_txt(path, filename, root_json)
    if(ask == 1 and system == 'Windows'):
        try:
            do_excel(path,filename,root_json)
            askUtoP = ask_url_to_pic()
            if askUtoP == 1:
                urlToPic.urlToPic.urlToPicWin(filename)
        except IOError as error:
            print(f'Error when generate Excel file:{error}')
    elif (ask == 1 and system == 'Darwin'):
        # try:
        do_excel_mac(path,filename,root_json)
        # except IOError as error:
        #     print(f'Error when generate Excel file:{error}')
    # pprint(response.text)
    ask2 = ask_excel('compare with old file')
    if(ask2 == 1):
        try:
            f1 = path+input(f'Enter first filename(older file): ')+'.txt'
            f2 = path+filename+'.txt'
            compare.compare_file(f1, f2)
        except IOError as error:
            print(f'Error when generate compared.txt file:{error}')
            print(f'Make sure u have the file existed and enter the right filename(without .txt)!!!')
            sys.exit()

if __name__ == '__main__':
    create_folder(path)
    if username == 'USERNAME or EMAIL':
        print(f'plz go config.py to set your USERNAME and PASSWORD')
        print(f'If you already have session.pkl you can only set USERNAME')
        os._exit(0)
    main()
