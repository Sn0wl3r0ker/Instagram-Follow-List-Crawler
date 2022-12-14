# from pprint import pprint
from bs4 import BeautifulSoup
import requests, pickle
import re
from config import username,password,headers,url,ajax_url,p_url,path
from datetime import datetime
from openpyxl import Workbook, load_workbook
import compare, urlToPic
import os, sys, time
import platform


def create_folder(path):
    if not os.path.isdir(path):
        os.makedirs(path)

def ask_excel(ask_option):
    flag = []
    yes_list = ['y','Y','yes']
    no_list = ['n','N','no','']
    while flag not in yes_list and flag not in no_list:
        flag = input(f'Do u want {ask_option}? y/n [n]: ')
        if(flag not in yes_list and flag not in no_list):
            print('plz enter y or n or ENTER!!!')
    # print(f'flag={flag}')
    if(flag in yes_list):
        return 1
    elif(flag in no_list):
        return 0
def ask_url_to_pic():
    ask3 = []
    yes_list = ['y','Y','yes','']
    no_list = ['n','N','no']
    while ask3 not in yes_list and ask3 not in no_list:
        ask3 = input(f'Do u want to transfer profile pic url to pics? It might *take more time*![y/n] [y]: ')
        if ask3 in yes_list:
            print(f'start getting pic files!!!')
            return 1
        elif ask3 in no_list:
            return 0
        else:
            print(f'plz enter y or n or empty to default!')

def do_excel(path,filename,root_json):              # 跑生成excel
    wb = Workbook()
    ws = wb.active
    title = ['username', 'full_name', 'profile_pic_url', 'Profile_Pic']
    ws.append(title)
    for users in root_json['users']:
        id = []
        id.append('@'+users['username'])
        id.append(users['full_name'])
        id.append(users['profile_pic_url']+'.jpg')
        ws.append(id)
    wb.save(path+filename+f'.xlsx')
    

def do_excel_mac(path,filename,root_json):
    if ask_url_to_pic() == 1:
        urlToPic.urlToPic.urlToPicinit(filename)
        time.sleep(3)
        print(f'waiting 3 seconds!')
        wb = load_workbook(os.path.abspath(path+filename+'pic.xlsm'), read_only=False, keep_vba=True)
        ws = wb.active
        title = ['username', 'full_name', 'profile_pic_url', 'Profile_Pic']
        ws.append(title)
        for users in root_json['users']:
            id = []
            id.append('@'+users['username'])
            id.append(users['full_name'])
            id.append(users['profile_pic_url']+'.jpg')
            ws.append(id)
        wb.save(path+filename+f'pic.xlsm')
        time.sleep(2)
        urlToPic.urlToPic.urlToPicMac(filename)
    else:
        do_excel(path,filename,root_json)



def do_txt(path,filename,root_json):                # 跑生成txt 
    i=1
    with open(path+filename+f'.txt', 'w+',encoding='utf-8') as f:
        for users in root_json['users']:
            # id = (f'{i}','@'+users['username'], users['full_name'])
            id = (f'@'+users['username'], users['full_name'])
            i+=1
        # reresponse = response.text.replace('\\u0026','&')
            f.write(str(id)+'\n')
        f.write(f'Total: {i-1} records!')
    print(f'Got {i-1} records!!!')

def main():
    system = platform.system()
    date = datetime.now().strftime("%Y%m%d-%H%M")
    time = int(datetime.now().timestamp())
    payload = {
        'username': f'{username}',
        'enc_password': f'#PWD_INSTAGRAM_BROWSER:0:{time}:{password}',
        'queryParams': {},
        'optIntoOneTap': 'false'
    }
    print(f'If target is private account, you have to follow it first!!!')
    while True:
        uid = str(input('Enter id: '))
        if (uid == ''):
            print(f'Do not leave blank!!!')
        else:
            break
    while True:
        opt_list = {'':'following','1':'following','2':'followers','following':'following','followers':'followers'}
        option = str(input('following[1]/followers[2] [1]: '))
        if(option in opt_list):
            option = opt_list[option]
            break
        else:
            print(f'enter 1 or 2 or following or followers!!!')
    try:
        fcount = int(input('Enter max num of following/followers [2000]: '))
    except ValueError:
        fcount = 2000

    opt_title = {
        'following': 'fwi',
        'followers': 'fwr',}
    filename = f'{uid}{date}{opt_title[option]}'
    ask = ask_excel('Excel file(will have profile pic)')
    # print(ask)
    with requests.session() as session:
        if not os.path.exists(f'{path}{username}session.pkl'):
            print('Getting sessions')      #session = requests.sess.....
            res = session.get(url)
            csrf = re.findall(r"csrf_token\":\"(.*?)\"",res.text)[0]
            cookies = res.cookies                   #res獲取第一次cookie和csrf
            cookies['csrf'] = csrf
            headers['x-csrftoken'] = csrf
            # print(headers)
            session.post(ajax_url, data=payload, headers=headers, cookies=cookies)  
            with open(f'{path}{username}session.pkl', 'wb') as f:
                pickle.dump(session.cookies, f)        #用現有cookie和csrf token 去取得登入的session
            headers['Referer'] = f'https://www.instagram.com/{uid}/following/'
        # print(req2.text)
        else:
            print('Reloading sessions and updating cookies')
            print('If can\'t get all the records, plz delete session.pkl file and try again later!')
            headers['Referer'] = f'https://www.instagram.com/{uid}/following/'
            with open(f'{path}{username}session.pkl', 'rb') as f:
                cookies = session.cookies.update(pickle.load(f))
                headers['x-csrftoken'] = session.cookies['csrftoken']
                # print(session.cookies)
                # print(headers)
        fsi=session.get(p_url+uid,cookies=cookies,headers=headers)    
        # print(fsi.text)
        
        try:
            # print(str(re.findall(r"id\":\"(.*?)\"",fsi.text)))
            friendid = str(re.findall(r"id\":\"(.*?)\"",fsi.text)[1])
            checkid = str(re.findall(r"id\":\"(.*?)\"",fsi.text)[-1])
            if(friendid == '236' or friendid == None or checkid == '236'):
                raise Exception
            print(f"userid:{friendid}")
        except:
            os.remove(f'{path}{username}session.pkl')
            print(f'error while checking userid')
            print(f'1.plz check the target username(no @)!!!')
            print(f'2.Make sure u set the right USERNAME and PASSWORD in *config.py* file!!!')
            print(f'3.your account might block by instagram server, plz try again later or change your ip!!')
            sys.exit()

        # url的後輟 可以像翻頁一樣去增加再爬取 或是直接爆max來爬取
        params = {
        'count': fcount,
        'max_id': '',
        'search_surface': 'follow_list_page'}
        response = session.get(f'https://i.instagram.com/api/v1/friendships/{friendid}/{option}/', params=params, cookies=cookies, headers=headers)
        # with open('debug.txt','w+', encoding='utf-8') as f:
        #     f.write(response.text)
        try:
            root_json = response.json()
        except requests.exceptions.JSONDecodeError as jsonError:
            print(f'Error when processing json file: {jsonError}')
            print(f'1.Make sure u set the right USERNAME and PASSWORD in *config.py* file!!!')
            print(f'2.your account might block by instagram server, plz try again later or change your ip!!')
            sys.exit()
            
    do_txt(path, filename, root_json)
    if(ask == 1 and system == 'Windows'):
        try:
            do_excel(path,filename,root_json)
            askUtoP = ask_url_to_pic()
            if askUtoP == 1:
                urlToPic.urlToPic.urlToPicWin(filename)
        except IOError as error:
            print(f'Error when generate Excel file:{error}')
    elif (ask == 1 and system == 'Darwin'):
        # try:
        do_excel_mac(path,filename,root_json)
        # except IOError as error:
        #     print(f'Error when generate Excel file:{error}')
    # pprint(response.text)
    ask2 = ask_excel('compare with old file')
    if(ask2 == 1):
        try:
            f1 = path+input(f'Enter first filename(older file): ')+'.txt'
            f2 = path+filename+'.txt'
            compare.compare_file(f1, f2)
        except IOError as error:
            print(f'Error when generate compared.txt file:{error}')
            print(f'Make sure u have the file existed and enter the right filename(without .txt)!!!')
            sys.exit()

if __name__ == '__main__':
    create_folder(path)
    if username == 'USERNAME or EMAIL':
        print(f'plz go config.py to set your USERNAME and PASSWORD')
        print(f'If you already have session.pkl you can only set USERNAME')
        os._exit(0)
    main()
