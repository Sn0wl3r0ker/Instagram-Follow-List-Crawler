import platform
from socket import timeout
from tkinter.filedialog import SaveAs
if platform.system() == 'Windows':
    import win32com.client as win32
import time, os
from config import vba_code
# import win32api
# import win32con
import xlwings as xw
import xlsxwriter
from openpyxl import workbook, load_workbook

class urlToPic:
    def urlToPicWin(filename):
        print(f'Generating new xlsm file~')
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(os.path.abspath('data/'+filename+'.xlsx'))
        excel.DisplayAlerts = False
        wb.DoNotPromptForConvert = True
        wb.CheckCompatibility = False
        wb.SaveAs(os.path.abspath(r"data/"+filename+"pic.xlsm"), FileFormat=52)
        excel.Application.Quit()
        del excel
        excel = None
        print('wait for 10 sec!')
        time.sleep(10)
        print('Start opening file and adding macro!\nIt might take more time! plz wait until finish!')
        excel2 = win32.gencache.EnsureDispatch('Excel.Application')
        excel2.Visible = False
        # key = win32api.RegOpenKeyEx(win32con.HKEY_CURRENT_USER,"SOFTWARE\\Microsoft\\Office\\16.0\\Excel\\Security", 0, win32con.KEY_ALL_ACCESS)
        # store original value
        # val, _ = win32api.RegQueryValueEx(key,'AccessVBOM') 
        # win32api.RegSetValueEx(key, "AccessVBOM", 0, win32con.REG_DWORD, 1)
        wb = excel2.Workbooks.Open(os.path.abspath(r'data/'+filename+'pic.xlsm'))
        origin_val = excel2.Application.AutomationSecurity
        excel2.Application.AutomationSecurity = 3
        mod = wb.VBProject.VBComponents.Add(1)
        mod.CodeModule.AddFromString(vba_code)
        # wb.VBProject.VBComponents.Import(os.path.abspath("vbaProject.bin"))
        print(f'start running macro!')
        excel2.Application.Run("Module1.URLPictureInsert")
        excel2.DisplayAlerts = False
        wb.DoNotPromptForConvert = True
        wb.CheckCompatibility = False
        excel2.Application.AutomationSecurity = origin_val
        wb.SaveAs(os.path.abspath(r"data/"+filename+"pic.xlsx"), FileFormat=51, ConflictResolution=2) # Macro disapears here
        print(f'Finished!!!')
        # win32api.RegCloseKey(key)
        excel2.Workbooks(1).Close(SaveChanges=1)
        excel2.Application.Quit()
        del excel2
        excel2 = None
        print(f'start removing xlsm file\n')
        os.remove(r'data/'+filename+'pic.xlsm')
        # return val

    def urlToPicinit(filename):
        workbook = xlsxwriter.Workbook(os.path.abspath('data/'+filename+'pic.xlsm'))
        workbook.add_vba_project(os.path.abspath('./vbaProject.bin'))
        workbook.close()

    def urlToPicMac(filename):
        # wb = load_workbook(os.path.abspath('data/'+filename2+'.xlsx'),read_only=False,keep_vba=True)
        # wb.save(os.path.abspath(r'data/'+filename2+'pic2.xlsm'))
        # book = app.books.open(os.path.abspath('data/'+filename2+'.xlsx'))
        # book.save()
        # wb = xw.Book(os.path.abspath('data/'+filename2+'.xlsx'))
        # wb.api.save_workbook_as(filename=os.path.abspath(r'data/'+filename2+'pic2.xlsx'),timeout=3000)
        # wb.close()
        # wb = None
        # print(f'wait for 3 sec')
        # time.sleep(3)
        # workbook = xlsxwriter.Workbook(os.path.abspath('data/'+filename+'pic.xlsm'))
        # workbook.add_vba_project(os.path.abspath('./vbaProject.bin'))
        # workbook.close()
        """
        app = xw.App(visible=True,add_book=False)
        app.api(timeout=6000)
        app.display_alerts =False
        app.screen_updating =False
        wb = app.books.open(os.path.abspath(r'data/'+filename+r'pic.xlsm'))
        # wb = app.books.open(os.path.abspath(r'data/'+filename+r'pic.xlsm'),timeout=6000)
        print(f'start running! plz wait!')
        time.sleep(3)
        macro = wb.macro('module1.URLPictureInsert')
        macro()"""

        app = xw.App(visible=True)
        # app.activate(steal_focus=True)
        # app.display_alerts =False
        # app.screen_updating =False
        # app.api(timeout=31)
        wb = app.books.open(os.path.abspath(r'data/'+filename+r'pic.xlsm'))
        print(f'start running! plz wait!')
        time.sleep(3)
        input(f'Press [enter] to start macro!!!: ')
        #解決 appscript time out 問題  run_VB_macro在Apple event timed out.報錯中找到
        #類似 https://github.com/xlwings/xlwings/issues/1955 問題，找api模組使用方法 win用pywin32 mac使用appscript!
        app.api.run_VB_macro(('module1.URLPictureInsert'),timeout=3000)
        # macro = wb.macro('module1.URLPictureInsert')
        # macro()
        # print(f'If you get the alert"Apple Event Timeout (-1712)" it\'s because there are too many pics to get and apple event timeout')
        # print(f'plz find aeosa/appscript/reference.py in your environment and set timeout 0 or 3000!!!')
        print(f'Excel will looks like freezing! It\'s normal on mac!!')
        input(f'Press [enter] to kill excel!!!: ')
        # print('3')
        # wb.save(os.path.abspath('data/'+filename2+'pic2.xlsm'))
        # print('4')
        
        print(f'finishing! wait for 5 seconds!!')
        time.sleep(5)
        wb.save()
        wb.close()
        app.kill()
        wb = None
        print(f'Finished! You have to convert xlsm to xlsx manually on mac!')
        # input(f'Press [enter] to continue!!!: ')
        # print(f'removing VBA code!')
        # time.sleep(5)
        
        # wb = load_workbook(os.path.abspath('data/'+filename+'pic.xlsm'),read_only=False,keep_vba=False)
        # wb.save(os.path.abspath(r'data/'+filename+'pic.xlsm'))
        # os.remove(os.path.abspath('data/'+filename+'pic.xlsm'))
